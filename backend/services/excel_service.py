import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from typing import List, Dict, Any

def generate_inquiry_excel(
    rows: List[Dict[str, Any]], 
    tpl: Dict[str, Any]
) -> bytes:
    """
    提供專業標楷體格式的詢價單生成服務。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "詢價表單"
    
    # --- 樣式定義 ---
    font_normal = Font(name="標楷體", size=12)
    font_bold = Font(name="標楷體", size=12, bold=True)
    font_blue = Font(name="標楷體", size=12, color="0000FF")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    border_thin = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    side_double = Side(style='double')
    
    # --- 表頭區 (Row 1-5) ---
    ws.merge_cells('A1:G1')
    ws['A1'] = tpl.get('company_name', '聖暉工程科技股份有限公司')
    ws['A1'].font = Font(name="標楷體", size=22, bold=True)
    ws['A1'].alignment = align_center
    
    ws.merge_cells('A2:G2')
    ws['A2'] = "投標-詢價單"
    ws['A2'].font = Font(name="標楷體", size=18, bold=True)
    ws['A2'].alignment = align_center
    
    ws.merge_cells('A3:G3')
    ws['A3'] = f"電話：{tpl.get('phone', '')}  傳真：{tpl.get('fax', '')}"
    ws['A3'].font = Font(name="標楷體", size=14)
    ws['A3'].alignment = align_center

    ws.merge_cells('A4:G4')
    ws['A4'] = f"地址：{tpl.get('address', '')}"
    ws['A4'].font = Font(name="標楷體", size=14)
    ws['A4'].alignment = align_center

    ws.merge_cells('A5:G5')
    ws['A5'] = f"Mail：{tpl.get('mail', '')}  聯絡人：{tpl.get('contact_person', '')}"
    ws['A5'].font = Font(name="標楷體", size=14)
    ws['A5'].alignment = align_center
    
    # 第 5 列底部加雙線
    for col in range(1, 8):
        ws.cell(row=5, column=col).border = Border(bottom=side_double)

    # --- 專案資訊 (Row 6-8) ---
    ws['A6'] = f"致：{tpl.get('vendor_to', '')}"
    ws.merge_cells('D6:G6')
    ws['D6'] = f"電話：{tpl.get('vendor_phone', '')}  傳真：{tpl.get('vendor_fax', '')}"
    ws['A6'].font = font_normal
    ws['D6'].font = font_normal
    
    ws.merge_cells('A7:G7')
    ws['A7'] = f"工程名稱：{tpl.get('project_name', '')}"
    ws['A7'].font = font_blue
    
    ws.merge_cells('A8:G8')
    ws['A8'] = f"工程地點：{tpl.get('project_location', '')}"
    ws['A8'].font = font_blue
    
    # --- 注意事項 (Row 9-11) ---
    ws['A9'] = "報價內容注意事項"
    ws['A9'].font = font_bold
    ws['A10'] = "1. 請報實售價，並於備註欄位註明報價廠牌、型號及折數(廠牌煩請備註)"
    ws['A10'].font = font_normal
    
    deadline_val = tpl.get('deadline', '')
    if deadline_val:
        if "懇請於" in deadline_val:
            ws['A11'] = f"2. {deadline_val}"
        else:
            ws['A11'] = f"2. 懇請於 {deadline_val} 前回覆報價"
    else:
        ws['A11'] = "2. 請參閱附件報價時間"
    ws['A11'].font = font_bold
    
    # --- 欄位標題 (Row 12) ---
    headers = ["項次", "品名/規格", "單位", "數量", "單價", "總價", "備註"]
    col_widths = [8, 45, 8, 12, 15, 15, 20]
    for col, (text, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=12, column=col)
        cell.value = text
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = border_thin
        ws.column_dimensions[chr(64 + col)].width = width
    
    # --- 填充內容 (Row 13+) ---
    for idx, r in enumerate(rows, 1):
        row_num = 12 + idx
        ws.cell(row=row_num, column=1, value=idx).border = border_thin
        ws.cell(row=row_num, column=2, value=r.get("description", "")).border = border_thin
        ws.cell(row=row_num, column=3, value=r.get("unit", "")).border = border_thin
        ws.cell(row=row_num, column=4, value=r.get("quantity", 0)).border = border_thin
        # 單價、總價、備註留白
        ws.cell(row=row_num, column=5, value="").border = border_thin
        ws.cell(row=row_num, column=6, value="").border = border_thin
        ws.cell(row=row_num, column=7, value="").border = border_thin
        
        # 設置對齊方式
        ws.cell(row=row_num, column=1).alignment = align_center
        ws.cell(row=row_num, column=2).alignment = align_left
        ws.cell(row=row_num, column=3).alignment = align_center
        ws.cell(row=row_num, column=4).alignment = align_center
        
        # 設置字體
        for col in range(1, 8):
            ws.cell(row=row_num, column=col).font = font_normal

    # --- 結尾：說明事項 ---
    current_row = 12 + len(rows) + 2
    ws.merge_cells(f'A{current_row}:G{current_row}')
    ws[f'A{current_row}'] = "註：本單不含稅。詳細內容請參閱標單規範。"
    ws[f'A{current_row}'].font = font_normal

    # 輸出成 Bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def generate_report_excel(
    lv1_data: Dict[str, Any], 
    lv2_data: Dict[str, Any], 
    project_name: str
) -> bytes:
    """
    生成專案成本報表 Excel，包含 LV.1 與 LV.2 兩個分頁。
    """
    wb = Workbook()
    
    # --- 樣式定義 ---
    font_header = Font(name="標楷體", size=16, bold=True)
    font_bold = Font(name="標楷體", size=11, bold=True)
    font_normal = Font(name="標楷體", size=11)
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    border_thin = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    def fill_sheet(ws, data, title):
        ws.title = title
        # 表頭
        ws.merge_cells('A1:G1')
        ws['A1'] = f"{project_name} - {title}"
        ws['A1'].font = font_header
        ws['A1'].alignment = align_center
        
        # 欄位標題
        headers = ["項次", "分類名稱", "廠商報價(元)", "占比", "公司成本(元)", "占比", "備註說明"]
        ws.append([]) # 空一行
        ws.append(headers)
        header_row = 3
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = font_bold
            cell.alignment = align_center
            cell.border = border_thin
            
        # 欄寬
        col_widths = [8, 30, 20, 10, 20, 10, 35]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64+i)].width = w
            
        # 數據內容
        categories = data.get("categories", [])
        total_internal = data.get("summary", {}).get("direct_internal", 0)
        
        for idx, cat in enumerate(categories, 1):
            row_num = header_row + idx
            ws.cell(row=row_num, column=1, value=idx).border = border_thin
            ws.cell(row=row_num, column=2, value=cat.get("name", "")).border = border_thin
            ws.cell(row=row_num, column=3, value=round(cat.get("supplier_total", 0))).border = border_thin
            ws.cell(row=row_num, column=4, value=cat.get("supplier_ratio", 0)/100).border = border_thin
            ws.cell(row=row_num, column=5, value=round(cat.get("internal_total", 0))).border = border_thin
            
            # 計算成本占比
            internal_ratio = (cat.get("internal_total", 0) / total_internal) if total_internal > 0 else 0
            ws.cell(row=row_num, column=6, value=internal_ratio).border = border_thin
            
            ws.cell(row=row_num, column=7, value=cat.get("remark", "")).border = border_thin
            
            # 格式化
            ws.cell(row=row_num, column=1).alignment = align_center
            ws.cell(row=row_num, column=2).alignment = align_center
            ws.cell(row=row_num, column=3).number_format = '#,##0'
            ws.cell(row=row_num, column=4).number_format = '0.0%'
            ws.cell(row=row_num, column=5).number_format = '#,##0'
            ws.cell(row=row_num, column=6).number_format = '0.0%'
            
            # 字體設定
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).font = font_normal

        # Summary 區塊
        summary = data.get("summary", {})
        start_summary = header_row + len(categories) + 2
        summary_items = [
            ("直接工程發包合計", summary.get("direct_supplier", 0), summary.get("direct_internal", 0)),
            (f"利潤及管理費 ({int(summary.get('profit_rate', 0.18)*100)}%)", summary.get("indirect_supplier", 0), summary.get("indirect_internal", 0)),
            ("小計", summary.get("direct_supplier", 0) + summary.get("indirect_supplier", 0), summary.get("direct_internal", 0) + summary.get("indirect_internal", 0)),
            (f"營業稅 ({int(summary.get('tax_rate', 0.05)*100)}%)", summary.get("tax_supplier", 0), 0),
            ("總計", summary.get("total_supplier", 0), summary.get("total_internal", 0))
        ]
        
        for i, (label, s_val, i_val) in enumerate(summary_items):
            r = start_summary + i
            ws.merge_cells(f'A{r}:B{r}')
            ws[f'A{r}'] = label
            ws[f'A{r}'].alignment = align_right
            ws[f'A{r}'].font = font_bold
            
            ws.cell(row=r, column=3, value=round(s_val)).font = font_bold
            ws.cell(row=r, column=3).number_format = '#,##0'
            
            ws.cell(row=r, column=5, value=round(i_val)).font = font_bold
            ws.cell(row=r, column=5).number_format = '#,##0'

    # 填充分頁一
    ws1 = wb.active
    fill_sheet(ws1, lv1_data, "大分類(LV.1)")
    
    # 填充分頁二
    ws2 = wb.create_sheet("成本管制(LV.2)")
    fill_sheet(ws2, lv2_data, "成本管制(LV.2)")
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
