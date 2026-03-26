from fastapi import FastAPI, UploadFile, File, Body, Request, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Dict, List, Any, Optional
import uvicorn
import pandas as pd
import io
from fastapi.responses import StreamingResponse
from services.excel_parser import ExcelParser
from services import project_service, category_service, excel_parser, diff_service, report_service
from database import engine, Base
import models

# 啟動時自動建立資料表 (如果不存在)
Base.metadata.create_all(bind=engine)

app = FastAPI()

@app.on_event("startup")
def startup_warmup():
    """伺服器啟動時預先載入快取，避免第一個請求等待大型 JSON 讀取"""
    if not project_service.USE_DB:
        try:
            project_service.load_projects()
            print(">>> [Startup] 專案快取預熱完成")
        except Exception as e:
            print(f">>> [Startup] 快取預熱失敗（可忽略）: {e}")


# 配置 CORS 中間件
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://localhost:3001",
        "http://localhost:3002",
        "https://project-cost-control-app.netlify.app"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class ProjectCreate(BaseModel):
    name: str
    client: str = ""
    location: str = ""
    manager: str = ""
    start_date: str = ""
    end_date: str = ""
    note: str = ""
    classification_depth: int = 2


class ProjectSettings(BaseModel):
    classification_depth: Optional[int] = None
    report_config: Optional[Dict[str, Any]] = None


class AnalyzeRequest(BaseModel):
    rows: List[Dict[str, Any]]
    max_depth: int = 2


class BatchClassifyRequest(BaseModel):
    row_indices: list[int]
    new_category_path: str
    add_keyword: bool = False
    keyword: str = ""
    classification_depth: int = 2

class ManualClassifyRequest(BaseModel):
    row_index: int
    new_category_path: str
    add_keyword: bool = False
    keyword: str = None
    item_name: str = None
    classification_depth: int = 2


@app.get("/")
def read_root():
    return {"status": "專案成本控制器後端已啟動", "message": "已連線至 8002", "version": "1.0.2"}


# === 專案管理 API ===

@app.get("/projects")
def list_projects():
    """取得所有專案列表（僅摘要，不含 rows 資料，大幅提升速度）"""
    return project_service.load_projects_summary()


@app.post("/projects")
def create_project(project: ProjectCreate):
    """建立新專案"""
    new_proj = project_service.create_project(
        name=project.name,
        client=project.client,
        location=project.location,
        manager=project.manager,
        start_date=project.start_date,
        end_date=project.end_date,
        note=project.note,
        classification_depth=project.classification_depth,
    )
    return {"status": "success", "project": new_proj}


@app.get("/projects/{project_id}")
def get_project(project_id: str, page: Optional[int] = None, page_size: int = 50, system_category: Optional[str] = None):
    """取得單一專案詳細資訊，支援分頁與類別過濾"""
    proj = project_service.get_project(project_id, page=page, page_size=page_size, system_category=system_category)
    if proj:
        return proj
    raise HTTPException(status_code=404, detail="找不到該專案")


@app.get("/projects/{project_id}/inquiry_rows")
def get_inquiry_rows(project_id: str, system_category: str):
    """僅取得特定分類的標單列資料 (極輕量化)"""
    return project_service.get_inquiry_rows(project_id, system_category)


@app.delete("/projects/{project_id}")
def delete_project(project_id: str):
    """刪除專案"""
    success = project_service.delete_project(project_id)
    if success:
        return {"status": "success"}
    raise HTTPException(status_code=404, detail="找不到該專案")


@app.patch("/projects/{project_id}/settings")
def update_project_settings(project_id: str, settings: ProjectSettings):
    """更新專案設定（例如分類深度、報表配置、詢價模板）"""
    # 將 settings 中非 None 的欄位提取出來
    update_dict = {k: v for k, v in settings.dict().items() if v is not None}
    
    updated = project_service.update_project_settings(project_id, update_dict)
    if updated:
        return {"status": "success", "project": updated}
    raise HTTPException(status_code=404, detail="找不到該專案")


# === 報表管理 API ===

@app.post("/projects/{project_id}/inquiry_rows/update")
async def update_inquiry_rows(project_id: str, payload: Dict[str, Any]):
    """批量更新特定類別下的詢價項目"""
    system_category = payload.get("system_category")
    updates = payload.get("updates", [])
    if not system_category:
        raise HTTPException(status_code=400, detail="必須提供類別名稱")
    
    success = project_service.update_inquiry_rows(project_id, system_category, updates)
    if success:
        return {"status": "success"}
    raise HTTPException(status_code=500, detail="更新失敗")

@app.get("/projects/{project_id}/reports")
def get_project_report(project_id: str, depth: int = 1):
    """取得專案報表數據 (LV.1 或 LV.2)"""
    return report_service.get_report_data(project_id, depth=depth)


@app.post("/projects/{project_id}/reports/config")
async def update_report_config(project_id: str, request: Request):
    """更新專案報表設定 (樓地板面積、各類別折數、利潤率等)"""
    body = await request.json()
    success = report_service.update_report_config(project_id, body)
    if success:
        return {"status": "success"}
    return {"status": "error", "message": "更新失敗"}


# === 分類管理 API ===

@app.post("/projects/{project_id}/compare")
async def compare_project_versions(project_id: str, request: Request):
    """比對兩個不同版本的標單檔案"""
    body = await request.json()
    base_idx = body.get("base_index", 0)  # 基準版本 (通常是舊的)
    target_idx = body.get("target_index", -1) # 目標版本 (通常是新的)
    
    project = project_service.get_project(project_id)
    if not project or not project.get("files"):
        return {"status": "error", "message": "專案或檔案不存在"}
    
    try:
        files = project["files"]
        if target_idx == -1: target_idx = len(files) - 1
        
        old_data = files[base_idx]["data"]["rows"]
        new_data = files[target_idx]["data"]["rows"]
        
        diff_result = diff_service.compare_and_merge(old_data, new_data)
        
        # 進行成本分析 (基於比對後的結果)
        analysis = category_service.analyze_project_data(diff_result["rows"], max_depth=project.get("classification_depth", 2))
        
        return {
            "status": "success", 
            "data": {
                "diff": diff_result,
                "analysis": analysis
            }
        }
    except Exception as e:
        print(f"[Error] Compare failed: {e}")
        return {"status": "error", "message": f"比對失敗: {str(e)}"}

@app.post("/projects/{project_id}/apply_diff")
async def apply_diff_version(project_id: str, request: Request):
    """將比對合併後的結果存為新版本"""
    body = await request.json()
    rows = body.get("rows")
    file_name = body.get("file_name", "合併版本")
    
    if not rows:
        return {"status": "error", "message": "無效的數據"}
        
    project = project_service.get_project(project_id)
    if not project:
        return {"status": "error", "message": "專案不存在"}
        
    # 分析並儲存
    analysis = category_service.analyze_project_data(rows, max_depth=project.get("classification_depth", 2))
    parsed_data = {
        "header_row": 0,
        "mapping": {}, # 合併版本可能沒有原始對應
        "rows": rows,
        "analysis": analysis
    }
    
    success = project_service.add_file_to_project(project_id, f"合併: {file_name}", parsed_data)
    if success:
        return {"status": "success", "message": "已儲存為新版本"}
    return {"status": "error", "message": "儲存失敗"}


@app.delete("/projects/{project_id}/files/{file_index}")
async def delete_project_version(project_id: str, file_index: int):
    """刪除特定版本的標單檔案"""
    success = project_service.delete_file_from_project(project_id, file_index)
    if success:
        return {"status": "success", "message": "已刪除該版本"}
    return {"status": "error", "message": "刪除失敗，找不到該版本"}


@app.get("/categories")
def get_categories():
    """取得機電系統分類配置"""
    return category_service.load_categories()


@app.post("/categories")
async def update_categories(request: Request):
    """更新機電系統分類配置"""
    print("--- 收到儲存請求 ---")
    try:
        categories = await request.json()
        print(f"資料大小: {len(str(categories))} 字元")
        category_service.save_categories(categories)
        return {"status": "success"}
    except Exception as e:
        print(f"Error saving categories: {e}")
        return {"status": "error", "message": str(e)}


@app.post("/analyze")
def analyze_rows(req: AnalyzeRequest):
    """手動重新分析數據列（用於切換深度時即時更新）"""
    try:
        analysis = category_service.analyze_project_data(req.rows, max_depth=req.max_depth)
        return {"status": "success", "analysis": analysis, "rows": req.rows}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.post("/projects/{project_id}/batch_classify")
def batch_classify(project_id: str, req: BatchClassifyRequest):
    """批次強制分類多個項目"""
    try:
        projects = project_service.load_projects()
        proj = next((p for p in projects if p["id"] == project_id), None)
        if not proj:
            return {"status": "error", "message": "找不到該專案"}
            
        files = proj.get("files", [])
        if not files:
            return {"status": "error", "message": "專案尚無檔案"}
            
        last_file = files[-1]
        data = last_file.get("data", {})
        rows = data.get("rows", [])
        
        updated_count = 0
        parts = req.new_category_path.split(" > ")
        short_path = " > ".join(parts[:req.classification_depth])
        
        for idx in req.row_indices:
            if 0 <= idx < len(rows):
                row = rows[idx]
                row["is_manual_category"] = True
                
                # 基於要求的邏輯：如果同時加入關鍵字，則把每個項目的描述作為子項目名稱
                item_description = row.get("description", "未命名項目")
                current_item_path = req.new_category_path
                
                if req.add_keyword:
                    # 建立品項子分類
                    current_item_path = f"{req.new_category_path} > {item_description}"
                    if req.keyword:
                        category_service.add_keyword_to_path(current_item_path, req.keyword)
                    else:
                        category_service.add_keyword_to_path(current_item_path)
                else:
                    # 不加入關鍵字，則純粹歸類到父分類
                    category_service.add_keyword_to_path(current_item_path)

                row["manual_raw_category"] = current_item_path
                row["system_category"] = " > ".join(current_item_path.split(" > ")[:req.classification_depth])
                updated_count += 1

        # 清除快取，觸發重新量分析
        category_service._categories_cache = None
        category_service._name_lookup_cache = None
        category_service._keyword_entries_cache = None
        category_service._classify_cache = {}
        
        analysis = category_service.analyze_project_data(rows, max_depth=req.classification_depth)
        
        data["analysis"] = analysis
        data["rows"] = rows
        
        # 使用新方法儲存，支援 JSON 與 DB
        project_service.update_project_files(project_id, files)
        
        return {"status": "success", "updated_count": updated_count, "data": data}
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return {"status": "error", "message": f"後端錯誤: {str(e)}"}


@app.post("/projects/{project_id}/manual_classify")
def manual_classify(project_id: str, req: ManualClassifyRequest):
    """手動強制分類單一項目，可選是否加入關鍵字學習"""
    try:
        projects = project_service.load_projects()
        proj = next((p for p in projects if p["id"] == project_id), None)
        if not proj:
            return {"status": "error", "message": "找不到該專案"}
            
        files = proj.get("files", [])
        if not files:
            return {"status": "error", "message": "專案尚無檔案"}
            
        last_file = files[-1]
        data = last_file.get("data", {})
        rows = data.get("rows", [])
        
        if req.row_index >= len(rows) or req.row_index < 0:
            return {"status": "error", "message": "無效的索引"}
            
        row = rows[req.row_index]
        row["is_manual_category"] = True
        row["manual_raw_category"] = req.new_category_path
        
        parts = req.new_category_path.split(" > ")
        row["system_category"] = " > ".join(parts[:req.classification_depth])

        # 如果有勾選學習且提供品項名稱，則建立品項子分類
        final_path = req.new_category_path
        if req.add_keyword and req.item_name:
            final_path = f"{req.new_category_path} > {req.item_name}"
            # 更新項目的顯示路徑為包含品項名稱的路徑（或維持原層級，取決於顯示需求，這裡選擇更新）
            # row["system_category"] = final_path 

        if req.add_keyword and req.keyword:
            category_service.add_keyword_to_path(final_path, req.keyword)
        else:
            category_service.add_keyword_to_path(final_path)

        category_service._categories_cache = None
        category_service._name_lookup_cache = None
        category_service._keyword_entries_cache = None
        category_service._classify_cache = {}
        
        analysis = category_service.analyze_project_data(rows, max_depth=req.classification_depth)
        
        data["analysis"] = analysis
        data["rows"] = rows
        
        # 使用新方法儲存，支援 JSON 與 DB
        project_service.update_project_files(project_id, files)
        
        return {"status": "success", "data": data}
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return {"status": "error", "message": f"後端錯誤: {str(e)}"}


# === 標單上傳 API ===

@app.post("/projects/{project_id}/upload")
async def upload_to_project(project_id: str, file: UploadFile = File(...)):
    """上傳標單至指定專案並自動儲存解析結果"""
    contents = await file.read()
    try:
        result = ExcelParser.parse_excel(contents)

        # 讀取該專案設定的分類深度
        proj = project_service.get_project(project_id)
        depth = proj.get("classification_depth", 3) if proj else 3

        # 依深度進行機電系統分類分析
        analysis = category_service.analyze_project_data(result["rows"], max_depth=depth)
        result["analysis"] = analysis

        # 診斷日誌
        print(f">>> [UPLOAD] Project: {project_id}, File: {file.filename}")
        print(f">>> [UPLOAD] Rows: {len(result.get('rows', []))}")

        success = project_service.add_file_to_project(project_id, file.filename, result)
        print(f">>> [UPLOAD] Save Success: {success}")

        return {"fileName": file.filename, "status": "success", "data": result}
    except Exception as e:
        import traceback
        print(f">>> [UPLOAD] Critical Error: {str(e)}")
        print(traceback.format_exc())
        return {"status": "error", "message": str(e)}


@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    """獨立上傳（不綁定專案，向後相容）"""
    contents = await file.read()
    try:
        result = ExcelParser.parse_excel(contents)
        return {
            "fileName": file.filename,
            "status": "success",
            "data": result
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.get("/debug/paths")
def debug_paths():
    from services.category_service import CATEGORIES_FILE, _categories_cache, _classify_cache
    import os
    return {
        "cwd": os.getcwd(),
        "categories_file": CATEGORIES_FILE,
        "abs_categories_file": os.path.abspath(CATEGORIES_FILE),
        "exists": os.path.exists(CATEGORIES_FILE),
        "cache_is_none": _categories_cache is None,
        "first_few_keys": list(_categories_cache.keys())[:5] if _categories_cache else [],
        "classify_cache_size": len(_classify_cache)
    }
@app.get("/projects/{project_id}/export")
def export_project_excel(project_id: str, version_idx: int = -1):
    """將專案內容匯出為 Excel 檔案，並將分類拆解為層級欄位 (LV.1, LV.2, LV.3)"""
    proj = project_service.get_project(project_id)
    if not proj or not proj.get("files"):
        return {"status": "error", "message": "找不到專案或尚未上傳標單"}
    
    # 決定要匯出的版本
    files = proj["files"]
    if version_idx == -1:
        version_idx = len(files) - 1
    
    if not (0 <= version_idx < len(files)):
        return {"status": "error", "message": f"無效的版本索引: {version_idx}"}
    
    target_file = files[version_idx]
    file_data = target_file.get("data", {})
    rows = file_data.get("rows", [])
    mapping = file_data.get("mapping", {})
    depth = proj.get("classification_depth", 3)
    
    if not rows:
        return {"status": "error", "message": "該版本無資料可匯出"}
    
    # 重新執行分析以確保最新分類結果以 rows 為準
    category_service.analyze_project_data(rows, max_depth=depth)
    
    # 決定要建立多少個層級欄位（至少 3 層，或根據專案設定更高）
    max_lv = max(3, depth)
    
    export_data = []
    for r in rows:
        new_row = {}
        # 1. 填入原始欄位 (依 mapping)
        for internal_key, chinese_name in mapping.items():
            new_row[chinese_name] = r.get(internal_key, "")
        
        # 2. 處理層級分類結果
        final_path = r.get("manual_raw_category") or r.get("system_category") or "未分類"
        path_parts = [p.strip() for p in final_path.split(" > ") if p.strip()]
        
        # 建立 LV.1 ~ LV.max_lv 欄位
        for i in range(1, max_lv + 1):
            col_name = f"LV.{i}"
            if i <= len(path_parts):
                new_row[col_name] = path_parts[i-1]
            else:
                new_row[col_name] = ""
                
        export_data.append(new_row)
        
    df_export = pd.DataFrame(export_data)
    
    # 將 DataFrame 寫入記憶體
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="成本控制分析")
    
    output.seek(0)
    
    filename = f"{proj['name']}_V{version_idx+1}_分級分析.xlsx"
    from urllib.parse import quote
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )


@app.get("/projects/{project_id}/inquiry_export")
async def export_inquiry_excel(project_id: str, category: str):
    """
    根據投標詢價範本格式，匯出指定類別的詢價 Excel。
    """
    # 修正：使用正確的函數名稱 get_project
    proj = project_service.get_project(project_id)
    if not proj:
        raise HTTPException(status_code=404, detail="Project not found")
        
    # 獲取模板設定，優先使用儲存的值，否則使用系統預設或專案內容
    raw_tpl = proj.get("report_config", {}).get("inquiry_template", {})
    tpl = {
        "company_name": raw_tpl.get("company_name") or "聖暉工程科技股份有限公司",
        "phone": raw_tpl.get("phone") or "02-2655-8067",
        "fax": raw_tpl.get("fax") or "02-2655-8073",
        "address": raw_tpl.get("address") or "115台北市南港區園區街3-2號5樓之2(軟體園區H棟)",
        "mail": raw_tpl.get("mail") or "yiyi_yang@acter.com.tw",
        "contact_person": raw_tpl.get("contact_person") or "楊尚嬑 小姐  分機:609",
        "project_name": raw_tpl.get("project_name") or proj.get("name", ""),
        "project_location": raw_tpl.get("project_location") or proj.get("location", ""),
        "deadline": raw_tpl.get("deadline") or ""
    }
    
    # 獲取該類別項目
    files = proj.get("files", [])
    if not files:
        raise HTTPException(status_code=400, detail="No versions found")
        
    latest_data = files[-1].get("data", {})
    rows = latest_data.get("rows", [])
    
    # 過濾屬於該類別及其子類別的項目
    # 注意：category 是 LV.2 或大類的顯示名稱
    filtered_rows = []
    for r in rows:
        cat_path = r.get("manual_raw_category") or r.get("system_category") or ""
        if category in cat_path:
            filtered_rows.append(r)
            
    # 使用 openpyxl 建立格式化 Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "詢價內容"
    
    # 字體與樣式 (依據使用者要求：標楷體，特定字重與大小)
    font_normal = Font(name="標楷體", size=12)
    font_bold = Font(name="標楷體", size=12, bold=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    border_thin = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # 填充表頭 (Row 1-11)
    # 填充表頭 (Row 1-5)
    ws.merge_cells('A1:G1')
    ws['A1'] = tpl.get('company_name', '聖暉工程科技股份有限公司')
    ws['A1'].font = Font(name="標楷體", size=22, bold=True)
    ws['A1'].alignment = align_center
    
    ws.merge_cells('A2:G2')
    ws['A2'] = "投標-詢價單"
    ws['A2'].font = Font(name="標楷體", size=18, bold=True)
    ws['A2'].alignment = align_center
    
    ws.merge_cells('A3:G3')
    ws['A3'] = f"電話：{tpl.get('phone', '')}  傳真：{tpl.get('fax', '')}"
    ws['A3'].font = Font(name="標楷體", size=14)
    ws['A3'].alignment = align_center

    ws.merge_cells('A4:G4')
    ws['A4'] = f"地址：{tpl.get('address', '')}"
    ws['A4'].font = Font(name="標楷體", size=14)
    ws['A4'].alignment = align_center

    ws.merge_cells('A5:G5')
    ws['A5'] = f"Mail：{tpl.get('mail', '')}  聯絡人：{tpl.get('contact_person', '')}"
    ws['A5'].font = Font(name="標楷體", size=14)
    ws['A5'].alignment = align_center
    
    # 分隔線 (Row 5 Bottom 使用雙線)
    side_double = Side(style='double')
    for col in range(1, 8):
        ws.cell(row=5, column=col).border = Border(bottom=side_double)

    # 專案資訊 (Row 6-8)
    ws['A6'] = "致："
    ws['D6'] = "電話："
    ws['F6'] = "傳真："
    ws['A6'].font = font_normal
    ws['D6'].font = font_normal
    ws['F6'].font = font_normal
    
    # 設置藍色字體給工程資訊 (Size 12)
    font_blue = Font(name="標楷體", size=12, color="0000FF")
    
    ws['A7'] = f"工程名稱：{tpl['project_name']}"
    ws['A7'].font = font_blue
    
    ws['A8'] = f"工程地點：{tpl['project_location']}"
    ws['A8'].font = font_blue
    
    ws['A9'] = "報價內容注意事項"
    ws['A9'].font = font_bold
    ws['A10'] = "1. 請報實售價，並於備註欄位註明報價廠牌、型號及折數(廠牌煩請備註)"
    ws['A10'].font = font_normal
    
    deadline_val = tpl['deadline']
    if deadline_val:
        # 如果使用者已經輸入了包含「懇請於」的完整句子，就不再重複加
        if "懇請於" in deadline_val:
            ws['A11'] = f"2. {deadline_val}"
        else:
            ws['A11'] = f"2. 懇請於 {deadline_val} 前回覆報價"
    else:
        ws['A11'] = "2. 請參閱附件報價時間"
    ws['A11'].font = font_bold
    
    # 欄位標題 (Row 12)
    headers = ["項次", "品名/規格", "單位", "數量", "單價", "總價", "備註"]
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=12, column=col)
        cell.value = text
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = border_thin
        
    # 填充內容 (Row 13+)
    for idx, r in enumerate(filtered_rows, 1):
        row_num = 12 + idx
        cells = []
        cells.append(ws.cell(row=row_num, column=1, value=idx))
        cells.append(ws.cell(row=row_num, column=2, value=r.get("description", "")))
        cells.append(ws.cell(row=row_num, column=3, value=r.get("unit", "")))
        cells.append(ws.cell(row=row_num, column=4, value=r.get("quantity", 0)))
        cells.append(ws.cell(row=row_num, column=5, value=None))
        cells.append(ws.cell(row=row_num, column=6, value=None))
        cells.append(ws.cell(row=row_num, column=7, value=r.get("remark", "")))
        
        for cell in cells:
            cell.border = border_thin
            cell.font = font_normal
        
        # 設置對齊
        ws.cell(row=row_num, column=1).alignment = align_center
        ws.cell(row=row_num, column=3).alignment = align_center
        ws.cell(row=row_num, column=4).alignment = align_right
        
    # 調整欄寬
    # 調整欄寬 (依據範本視覺比例)
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    safe_category = category.replace("/", "_").replace("\\", "_")
    filename = f"詢價單_{safe_category}.xlsx"
    from urllib.parse import quote
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )


if __name__ == "__main__":
    # 強制指定運行於 8002 埠，並開啟熱重載以套用後續修正
    uvicorn.run("main:app", host="0.0.0.0", port=8002, reload=True)